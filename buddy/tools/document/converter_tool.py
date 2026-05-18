from __future__ import annotations

import csv
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Union

from buddy.prompts.converter_prompts import CONVERTER_TOOL_PROMPT

_TOOL = "converter"

# ── support matrix ────────────────────────────────────────────────────────────

_IMAGE_EXTS: Set[str] = {
    ".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".tiff", ".tif",
}

# source_ext → set of valid destination_formats
_SUPPORT: Dict[str, Set[str]] = {
    ".docx": {"pdf", "html"},
    ".xlsx": {"pdf", "html", "csv"},
    ".pptx": {"pdf", "png", "jpg"},
    ".odt":  {"pdf", "docx", "html"},
    ".odp":  {"pdf", "png"},
    ".ods":  {"pdf", "csv"},
    ".pdf":  {"txt", "png", "jpg"},
    ".html": {"pdf", "docx"},
    ".htm":  {"pdf", "docx"},
    ".md":   {"html", "pdf", "docx"},
    ".txt":  {"pdf", "docx"},
    **{ext: {"pdf", "png", "jpg", "bmp", "webp"} for ext in _IMAGE_EXTS},
}

# (src_ext, dest_fmt) → backend key
_BACKEND: Dict[tuple, str] = {
    (".docx", "pdf"): "lo",  (".docx", "html"): "lo",
    (".xlsx", "pdf"): "lo",  (".xlsx", "html"): "lo",  (".xlsx", "csv"): "xlsx_csv",
    (".pptx", "pdf"): "lo",  (".pptx", "png"):  "lo",  (".pptx", "jpg"):  "lo",
    (".odt",  "pdf"): "lo",  (".odt",  "docx"): "lo",  (".odt",  "html"): "lo",
    (".odp",  "pdf"): "lo",  (".odp",  "png"):  "lo",
    (".ods",  "pdf"): "lo",  (".ods",  "csv"):  "lo",
    (".html", "pdf"): "lo",  (".html", "docx"): "lo",
    (".htm",  "pdf"): "lo",  (".htm",  "docx"): "lo",
    (".txt",  "pdf"): "lo",  (".txt",  "docx"): "lo",
    (".pdf",  "txt"): "pdf_txt",
    (".pdf",  "png"): "pdf_img",
    (".pdf",  "jpg"): "pdf_img",
    (".md",   "html"): "md_html",
    (".md",   "pdf"):  "md_lo",
    (".md",   "docx"): "md_lo",
}
# image→image and image→pdf handled dynamically in _route()


# ── helpers ───────────────────────────────────────────────────────────────────

def _resolve(raw: str) -> str:
    import os
    p = os.path.expanduser(os.path.expandvars(raw.strip()))
    if not os.path.isabs(p):
        p = os.path.join(os.path.expanduser("~"), p)
    return p


def _ok(**kw: Any) -> Dict[str, Any]:
    return {"STATUS": "success", **kw}


def _err(msg: str, **kw: Any) -> Dict[str, Any]:
    return {"STATUS": "failed", "TOOL": _TOOL, "ERROR": msg, **kw}


def _parse_range(range_str: str, total: int) -> Optional[Set[int]]:
    s = str(range_str or "all").strip().lower()
    if s == "all":
        return None
    try:
        if "-" in s:
            a, b = s.split("-", 1)
            return set(range(int(a), int(b) + 1))
        return {int(s)}
    except (ValueError, IndexError):
        return None


def _filter_files_by_range(
    files: List[Path], range_str: str,
) -> List[Path]:
    keep = _parse_range(range_str, len(files))
    if keep is None:
        return files
    return [f for i, f in enumerate(files, 1) if i in keep]


def _route(src_ext: str, dest_fmt: str) -> Optional[str]:
    if src_ext in _IMAGE_EXTS:
        if dest_fmt == "pdf":
            return "img_pdf"
        if dest_fmt in {"png", "jpg", "jpeg", "bmp", "gif", "webp"}:
            return "img_img"
        return None
    return _BACKEND.get((src_ext, dest_fmt))


def _lo_binary() -> Optional[str]:
    # On macOS, LibreOffice is often not in PATH, so check default install location
    import sys
    import os
    if sys.platform == "darwin":
        mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        if os.path.exists(mac_path):
            return mac_path
    return shutil.which("soffice") or shutil.which("libreoffice")

def _lo_missing_error() -> Dict[str, Any]:
    import sys
    if sys.platform == "darwin":
        install_cmd = "brew install --cask libreoffice"
    elif sys.platform.startswith("linux"):
        install_cmd = "sudo apt-get update && sudo apt-get install libreoffice"
    else:
        install_cmd = "Download from https://www.libreoffice.org/"
    return _err(
        "LIBREOFFICE_MISSING: LibreOffice is required for this conversion but is not installed or not in PATH.\n"
        f"Action required: {install_cmd}"
    )


def _run_lo(
    lo: str, src: Path, dest_fmt: str, output_dir: Path,
) -> Dict[str, Any]:
    try:
        result = subprocess.run(
            [lo, "--headless", "--convert-to", dest_fmt,
             "--outdir", str(output_dir), str(src)],
            capture_output=True, text=True, timeout=120,
        )
    except subprocess.TimeoutExpired:
        return _err("CONVERSION_FAILED: LibreOffice timed out after 120 seconds")
    except Exception as e:
        return _err(f"CONVERSION_FAILED: could not launch LibreOffice: {e}")

    if result.returncode != 0:
        msg = result.stderr.strip() or result.stdout.strip()
        return _err(
            f"CONVERSION_FAILED: LibreOffice exited {result.returncode}: {msg}",
            LO_OUTPUT=msg,
        )
    return _ok()


# ── tool ──────────────────────────────────────────────────────────────────────

class ConverterTool:
    tool_name = _TOOL
    version = "1.0.0"

    def get_info(self) -> Dict[str, Any]:
        return {
            "name": self.tool_name,
            "version": self.version,
            "description": (
                "WHEN: converting any file from one format to another.\n\n"
                "FUNCTION:\n"
                "  convert(source, destination_format, output_dir, slide_range?, page_range?, sheet?)\n\n"
                "SUPPORTED:\n"
                "  docx → pdf html\n"
                "  xlsx → pdf html csv\n"
                "  pptx → pdf png jpg\n"
                "  pdf  → txt png jpg\n"
                "  html/md/txt → pdf docx\n"
                "  images → pdf (single or array merged), cross-format (png↔jpg↔bmp↔webp)\n\n"
                "NOT: file content creation/editing → use word/excel/ppt/pdf tools"
            ),
            "prompt": CONVERTER_TOOL_PROMPT,
        }

    async def execute(
        self,
        function: str,
        arguments: Dict[str, Any],
        on_progress: Optional[Callable] = None,
        goal: str = "",
        brain: Any = None,
        **_: Any,
    ) -> Dict[str, Any]:
        fn = str(function or "").strip().lower()
        if fn == "convert":
            return self._convert(arguments)
        return _err(f"Unknown function: {function!r}. Must be: convert")

    # ── convert ───────────────────────────────────────────────────────────────

    def _convert(self, args: Dict[str, Any]) -> Dict[str, Any]:
        source_raw = args.get("source")
        dest_fmt = str(args.get("destination_format") or "").strip().lower().lstrip(".")
        raw_dir = str(args.get("output_dir") or "").strip()

        if source_raw is None:
            return _err("'source' is required — absolute path or array of image paths")
        if not dest_fmt:
            return _err("'destination_format' is required — e.g. 'pdf', 'png', 'docx'")
        if not raw_dir:
            return _err("'output_dir' is required — absolute path to output directory")

        output_dir = Path(_resolve(raw_dir))
        is_array = isinstance(source_raw, list)

        # ── array source: images → pdf only ──────────────────────────────────
        if is_array:
            if dest_fmt != "pdf":
                return _err("Array source only supports destination_format='pdf' (merge images into PDF)")
            if not source_raw:
                return _err("'source' array must not be empty")
            sources = [Path(_resolve(str(s))) for s in source_raw]
            missing = [str(s) for s in sources if not s.exists()]
            if missing:
                return _err("FILE_NOT_FOUND: one or more source files not found", MISSING=missing)
            try:
                output_dir.mkdir(parents=True, exist_ok=True)
            except PermissionError:
                return _err(f"PERMISSION_DENIED: cannot create output_dir: {output_dir}")
            return self._images_to_pdf(sources, output_dir)

        # ── single source ─────────────────────────────────────────────────────
        src = Path(_resolve(str(source_raw)))
        if not src.exists():
            return _err(f"FILE_NOT_FOUND: {src} — use fs_browse.find to locate it")
        if not src.is_file():
            return _err(f"FILE_NOT_FOUND: source is a directory, not a file: {src}")

        src_ext = src.suffix.lower()

        # validate
        valid_fmts = _SUPPORT.get(src_ext)
        if valid_fmts is None:
            return _err(
                f"UNSUPPORTED: source format '{src_ext}' is not supported.",
                SUPPORTED_SOURCE_FORMATS=sorted(_SUPPORT.keys()),
            )
        if dest_fmt not in valid_fmts:
            return _err(
                f"UNSUPPORTED: '{src_ext}' → '{dest_fmt}' is not a supported combination.",
                SUPPORTED_FORMATS=sorted(valid_fmts),
            )

        try:
            output_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            return _err(f"PERMISSION_DENIED: cannot create output_dir: {output_dir}")

        backend = _route(src_ext, dest_fmt)
        if backend is None:
            return _err(f"UNSUPPORTED: no backend for '{src_ext}' → '{dest_fmt}'")

        slide_range = str(args.get("slide_range") or "all")
        page_range = str(args.get("page_range") or "all")
        sheet = args.get("sheet")

        try:
            if backend == "lo":
                return self._via_libreoffice(src, dest_fmt, output_dir, slide_range)
            if backend == "xlsx_csv":
                return self._xlsx_to_csv(src, output_dir, sheet)
            if backend == "pdf_txt":
                return self._pdf_to_txt(src, output_dir, page_range)
            if backend == "pdf_img":
                return self._pdf_to_images(src, dest_fmt, output_dir, page_range)
            if backend == "md_html":
                return self._md_to_html(src, output_dir)
            if backend == "md_lo":
                return self._md_via_lo(src, dest_fmt, output_dir)
            if backend == "img_pdf":
                return self._images_to_pdf([src], output_dir)
            if backend == "img_img":
                return self._image_convert(src, dest_fmt, output_dir)
        except PermissionError:
            return _err(f"PERMISSION_DENIED: cannot write to {output_dir}")
        except Exception as e:
            return _err(f"Conversion failed: {e}")

        return _err(f"Unhandled backend: {backend}")

    # ── backends ──────────────────────────────────────────────────────────────

    def _via_libreoffice(
        self, src: Path, dest_fmt: str, output_dir: Path, range_str: str,
    ) -> Dict[str, Any]:
        lo = _lo_binary()
        if not lo:
            return _lo_missing_error()

        result = _run_lo(lo, src, dest_fmt, output_dir)
        if result["STATUS"] == "failed":
            return result

        # Collect output files (LibreOffice names them {stem}.{fmt} or {stem}-001.{fmt} for slides)
        files = sorted(output_dir.glob(f"{src.stem}*.{dest_fmt}"))
        if not files:
            # Some LO versions use different naming — glob broadly
            files = sorted(output_dir.glob(f"*{src.stem}*.{dest_fmt}"))
        if not files:
            return _err(
                f"LibreOffice reported success but no .{dest_fmt} files found in {output_dir}"
            )

        # Apply slide_range filter for multi-file image outputs
        if dest_fmt in {"png", "jpg"} and range_str != "all":
            files = _filter_files_by_range(files, range_str)

        return _ok(
            SOURCE=str(src),
            FILES=[str(f) for f in files],
            OUTPUT_DIR=str(output_dir),
            FILE_COUNT=len(files),
        )

    def _xlsx_to_csv(
        self, src: Path, output_dir: Path, sheet: Optional[str],
    ) -> Dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            return _err("IMPORT_ERROR: openpyxl not installed. Run: pip install openpyxl")

        wb = openpyxl.load_workbook(str(src), data_only=True)
        try:
            if sheet:
                if sheet not in wb.sheetnames:
                    return _err(
                        f"Sheet '{sheet}' not found.",
                        AVAILABLE_SHEETS=wb.sheetnames,
                    )
                ws = wb[sheet]
            else:
                ws = wb.active

            out = output_dir / f"{src.stem}.csv"
            with open(str(out), "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else str(v) for v in row])

            return _ok(
                SOURCE=str(src),
                FILES=[str(out)],
                OUTPUT_DIR=str(output_dir),
                SHEET=ws.title,
                FILE_COUNT=1,
            )
        finally:
            wb.close()

    def _pdf_to_txt(
        self, src: Path, output_dir: Path, page_range: str,
    ) -> Dict[str, Any]:
        try:
            import pdfplumber
        except ImportError:
            return _err("IMPORT_ERROR: pdfplumber not installed. Run: pip install pdfplumber")

        with pdfplumber.open(str(src)) as pdf:
            total = len(pdf.pages)
            keep = _parse_range(page_range, total)
            pages = [p for i, p in enumerate(pdf.pages, 1) if keep is None or i in keep]
            text = "\n\n".join(p.extract_text() or "" for p in pages)

        out = output_dir / f"{src.stem}.txt"
        out.write_text(text, encoding="utf-8")

        return _ok(
            SOURCE=str(src),
            FILES=[str(out)],
            OUTPUT_DIR=str(output_dir),
            PAGES_EXTRACTED=len(pages),
            FILE_COUNT=1,
        )

    def _pdf_to_images(
        self, src: Path, dest_fmt: str, output_dir: Path, page_range: str,
    ) -> Dict[str, Any]:
        try:
            import pypdfium2 as pdfium
        except ImportError:
            return _err("IMPORT_ERROR: pypdfium2 not installed. Run: pip install pypdfium2")

        save_fmt = "JPEG" if dest_fmt == "jpg" else dest_fmt.upper()
        ext = "jpg" if dest_fmt == "jpg" else dest_fmt

        files: List[Path] = []
        with pdfium.PdfDocument(str(src)) as pdf:
            total = len(pdf)
            keep = _parse_range(page_range, total)
            for page_num in range(total):
                if keep is not None and (page_num + 1) not in keep:
                    continue
                page = pdf[page_num]
                try:
                    bitmap = page.render(scale=2)
                    pil_img = bitmap.to_pil()
                    bitmap.close()
                finally:
                    page.close()
                if dest_fmt == "jpg":
                    pil_img = pil_img.convert("RGB")
                out = output_dir / f"{src.stem}-{page_num + 1:03d}.{ext}"
                pil_img.save(str(out), save_fmt)
                files.append(out)

        return _ok(
            SOURCE=str(src),
            FILES=[str(f) for f in files],
            OUTPUT_DIR=str(output_dir),
            FILE_COUNT=len(files),
        )

    def _md_to_html(self, src: Path, output_dir: Path) -> Dict[str, Any]:
        try:
            import markdown as md_lib
        except ImportError:
            return _err("IMPORT_ERROR: markdown not installed. Run: pip install Markdown")

        text = src.read_text(encoding="utf-8", errors="replace")
        html = md_lib.markdown(text)
        out = output_dir / f"{src.stem}.html"
        out.write_text(html, encoding="utf-8")

        return _ok(
            SOURCE=str(src),
            FILES=[str(out)],
            OUTPUT_DIR=str(output_dir),
            FILE_COUNT=1,
        )

    def _md_via_lo(
        self, src: Path, dest_fmt: str, output_dir: Path,
    ) -> Dict[str, Any]:
        lo = _lo_binary()
        if not lo:
            return _lo_missing_error()

        try:
            import markdown as md_lib
        except ImportError:
            return _err("IMPORT_ERROR: markdown not installed. Run: pip install Markdown")

        # md → temp html → LibreOffice → dest_fmt
        text = src.read_text(encoding="utf-8", errors="replace")
        html = f"<html><body>{md_lib.markdown(text)}</body></html>"

        with tempfile.NamedTemporaryFile(
            suffix=".html", delete=False, mode="w", encoding="utf-8"
        ) as tmp:
            tmp.write(html)
            tmp_path = Path(tmp.name)

        try:
            result = _run_lo(lo, tmp_path, dest_fmt, output_dir)
            if result["STATUS"] == "failed":
                return result

            # LibreOffice names it after the temp file stem — rename to source stem
            generated = list(output_dir.glob(f"{tmp_path.stem}.{dest_fmt}"))
            if generated:
                final = output_dir / f"{src.stem}.{dest_fmt}"
                shutil.move(str(generated[0]), str(final))
            else:
                final_candidates = list(output_dir.glob(f"*.{dest_fmt}"))
                if not final_candidates:
                    return _err(f"Conversion succeeded but no .{dest_fmt} file found in {output_dir}")
                final = final_candidates[0]

            return _ok(
                SOURCE=str(src),
                FILES=[str(final)],
                OUTPUT_DIR=str(output_dir),
                FILE_COUNT=1,
            )
        finally:
            tmp_path.unlink(missing_ok=True)

    def _images_to_pdf(self, sources: List[Path], output_dir: Path) -> Dict[str, Any]:
        try:
            from PIL import Image
        except ImportError:
            return _err("IMPORT_ERROR: Pillow not installed. Run: pip install Pillow")

        images: List[Any] = []
        for s in sources:
            img = Image.open(str(s))
            if img.mode in ("RGBA", "P", "LA"):
                img = img.convert("RGB")
            elif img.mode != "RGB":
                img = img.convert("RGB")
            images.append(img)

        out = output_dir / f"{sources[0].stem}.pdf"
        if len(images) == 1:
            images[0].save(str(out), "PDF", resolution=150.0)
        else:
            images[0].save(
                str(out), "PDF", resolution=150.0,
                save_all=True, append_images=images[1:],
            )

        return _ok(
            SOURCES=[str(s) for s in sources],
            FILES=[str(out)],
            OUTPUT_DIR=str(output_dir),
            PAGES=len(images),
            FILE_COUNT=1,
        )

    def _image_convert(
        self, src: Path, dest_fmt: str, output_dir: Path,
    ) -> Dict[str, Any]:
        try:
            from PIL import Image
        except ImportError:
            return _err("IMPORT_ERROR: Pillow not installed. Run: pip install Pillow")

        img = Image.open(str(src))
        ext = "jpg" if dest_fmt == "jpg" else dest_fmt
        save_fmt = "JPEG" if dest_fmt == "jpg" else dest_fmt.upper()

        if dest_fmt == "jpg" and img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")

        out = output_dir / f"{src.stem}.{ext}"
        img.save(str(out), save_fmt)

        return _ok(
            SOURCE=str(src),
            FILES=[str(out)],
            OUTPUT_DIR=str(output_dir),
            FILE_COUNT=1,
        )


# ── registry ──────────────────────────────────────────────────────────────────

TOOL_NAME = "converter"
TOOL_CLASS = ConverterTool


def get_tool() -> ConverterTool:
    return ConverterTool()
