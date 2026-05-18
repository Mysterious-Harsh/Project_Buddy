from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from buddy.prompts.excel_prompts import EXCEL_TOOL_PROMPT

_TOOL = "excel"
_PREVIEW_ROWS = 5


# ── helpers ──────────────────────────────────────────────────────────────────

def _resolve(raw: str) -> str:
    import os
    p = os.path.expanduser(os.path.expandvars(raw.strip()))
    if not os.path.isabs(p):
        p = os.path.join(os.path.expanduser("~"), p)
    return p


def _ok(**kw: Any) -> Dict[str, Any]:
    return {"STATUS": "success", **kw}


def _err(msg: str, **kw: Any) -> Dict[str, Any]:
    return {"STATUS": "failed", "TOOL": _TOOL, "ERROR": msg, **kw}


def _needs_confirm(preview: str, **kw: Any) -> Dict[str, Any]:
    return {
        "STATUS": "failed", "TOOL": _TOOL,
        "NEEDS_CONFIRMATION": True,
        "PREVIEW": preview,
        "NOTE": "Call again with confirmed=true after user approves.",
        **kw,
    }


def _col_letter(col_idx: int) -> str:
    """1-based column index → Excel letter(s). 1→A, 26→Z, 27→AA."""
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


# ── tool ─────────────────────────────────────────────────────────────────────

class ExcelTool:
    tool_name = _TOOL
    version = "1.0.0"

    def get_info(self) -> Dict[str, Any]:
        return {
            "name": self.tool_name,
            "version": self.version,
            "description": (
                "WHEN: any operation on an .xlsx spreadsheet — data entry, cell editing, sheet management, searching values.\n\n"
                "FUNCTIONS:\n"
                "  create(path, sheets[])                    — new workbook; sheets: [{name, headers[], rows[[]]}]\n"
                "  read(path, sheet?, preview_rows?)          — returns sheet names, headers, row count, data preview\n"
                "  edit(path, sheet, operations[])            — atomic batch edits: set_cell, add_row, insert_row, delete_row, add_sheet, delete_sheet, rename_sheet\n"
                "  search(path, sheet?, query | col+op+val)   — find rows by text match OR column condition (= != > < >= <= contains starts_with ends_with)\n\n"
                "CHAIN: always call read before edit to confirm exact sheet names and column headers (case-sensitive). "
                "search returns row indices → feed into edit to target specific rows.\n"
                "NOT: .docx → word | .pdf → pdf | .csv/.txt plain reads → fs_read | .csv/.txt create/edit → fs_write | export/convert → converter"
            ),
            "prompt": EXCEL_TOOL_PROMPT,
        }

    async def execute(
        self,
        function: str,
        arguments: Dict[str, Any],
        on_progress: Optional[Callable] = None,
        goal: str = "",
        brain: Any = None,
        **_: Any,
    ) -> Dict[str, Any]:
        fn = str(function or "").strip().lower()
        if fn == "create":
            return self._create(arguments)
        if fn == "read":
            return self._read(arguments)
        if fn == "edit":
            return self._edit(arguments)
        if fn == "search":
            return self._search(arguments)
        return _err(
            f"Unknown function: {function!r}. Must be: create, read, edit, search"
        )

    # ── create ───────────────────────────────────────────────────────────────

    def _create(self, args: Dict[str, Any]) -> Dict[str, Any]:
        raw = str(args.get("path") or "").strip()
        if not raw:
            return _err("excel.create: 'path' is required — provide an absolute .xlsx path")
        if not raw.lower().endswith(".xlsx"):
            return _err("excel.create: path must end in .xlsx")

        sheets_data = args.get("sheets")
        if not sheets_data or not isinstance(sheets_data, list):
            return _err("sheets is required and must be a non-empty list")

        path = _resolve(raw)
        p = Path(path)

        if p.exists() and not args.get("confirmed"):
            return _needs_confirm(
                f"File already exists: {path}\n"
                "Setting confirmed=true will overwrite it — all existing sheets will be lost."
            )

        try:
            import openpyxl
        except ImportError:
            return _err("openpyxl not installed. Run: pip install openpyxl")

        try:
            wb = openpyxl.Workbook()
            default_ws = wb.active
            if default_ws is None:
                default_ws = wb.create_sheet("Sheet1")
            total_rows = 0

            for i, sheet in enumerate(sheets_data):
                if not isinstance(sheet, dict):
                    return _err(f"sheets[{i}] must be an object with name, headers, rows")

                name = str(sheet.get("name") or f"Sheet{i + 1}").strip()
                headers: List = sheet.get("headers") or []
                rows: List = sheet.get("rows") or []

                if i == 0:
                    ws = default_ws
                    ws.title = name
                else:
                    if name in wb.sheetnames:
                        return _err(f"Duplicate sheet name: '{name}'. All sheet names must be unique.")
                    ws = wb.create_sheet(name)

                if headers:
                    ws.append([str(h) for h in headers])
                for row in rows:
                    ws.append(list(row))
                    total_rows += 1

            p.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(p))
            return _ok(PATH=str(p), SHEETS=len(sheets_data), ROWS_WRITTEN=total_rows)

        except PermissionError:
            return _err(f"Permission denied: {path}")
        except Exception as e:
            return _err(f"Failed to create workbook: {e}")

    # ── read ─────────────────────────────────────────────────────────────────

    def _read(self, args: Dict[str, Any]) -> Dict[str, Any]:
        raw = str(args.get("path") or "").strip()
        if not raw:
            return _err("excel.read: 'path' is required — provide an absolute .xlsx path")

        path = _resolve(raw)
        p = Path(path)
        if not p.exists():
            return _err(f"excel.read: file not found: {path} — use fs_browse.find to locate it")
        if not p.is_file():
            return _err(f"excel.read: path is a directory, not a file: {path} — provide path to an .xlsx file")

        preview_rows = max(1, int(args.get("preview_rows") or _PREVIEW_ROWS))
        sheet_name: Optional[str] = args.get("sheet")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), data_only=True)

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return _err(
                        f"Sheet '{sheet_name}' not found.",
                        AVAILABLE_SHEETS=wb.sheetnames,
                    )
                detail = self._sheet_detail(wb[sheet_name], preview_rows)
                return _ok(PATH=str(p), SHEET_DETAIL=detail)

            summary = [
                {
                    "name": name,
                    "row_count": max(0, (wb[name].max_row or 1) - 1),
                    "column_count": wb[name].max_column or 0,
                }
                for name in wb.sheetnames
            ]
            active = wb.active or wb[wb.sheetnames[0]]
            return _ok(
                PATH=str(p),
                SHEET_COUNT=len(wb.sheetnames),
                SHEETS=summary,
                ACTIVE_SHEET=self._sheet_detail(active, preview_rows),
            )

        except ImportError:
            return self._read_pandas_fallback(path, sheet_name, preview_rows)
        except PermissionError:
            return _err(f"Permission denied: {path}")
        except Exception as e:
            return _err(f"Failed to read workbook: {e}")

    def _sheet_detail(self, ws: Any, preview_rows: int) -> Dict[str, Any]:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return {"name": ws.title, "headers": [], "row_count": 0, "column_count": 0, "preview": []}

        headers = [str(h) if h is not None else "" for h in all_rows[0]]
        data_rows = all_rows[1:]
        preview = [list(r) for r in data_rows[:preview_rows]]

        return {
            "name": ws.title,
            "headers": headers,
            "row_count": len(data_rows),
            "column_count": len(headers),
            "preview": preview,
        }

    def _read_pandas_fallback(
        self, path: str, sheet_name: Optional[str], preview_rows: int
    ) -> Dict[str, Any]:
        try:
            import pandas as pd
            xl = pd.ExcelFile(path)
            names = xl.sheet_names
            target = sheet_name if (sheet_name and sheet_name in names) else names[0]
            df = xl.parse(target)
            preview = df.head(preview_rows).values.tolist()
            return _ok(
                PATH=path,
                SHEET_COUNT=len(names),
                SHEETS=[{"name": n} for n in names],
                ACTIVE_SHEET={
                    "name": target,
                    "headers": list(df.columns),
                    "row_count": len(df),
                    "column_count": len(df.columns),
                    "preview": preview,
                },
                NOTE=(
                    "Read via pandas fallback — openpyxl not installed. "
                    "Install for full feature support: pip install openpyxl"
                ),
            )
        except ImportError:
            return _err(
                "Neither openpyxl nor pandas is installed. Run: pip install openpyxl"
            )
        except Exception as e:
            return _err(f"Fallback read failed: {e}")

    # ── edit ─────────────────────────────────────────────────────────────────

    def _edit(self, args: Dict[str, Any]) -> Dict[str, Any]:
        raw = str(args.get("path") or "").strip()
        if not raw:
            return _err("excel.edit: 'path' is required — provide an absolute .xlsx path")

        operations = args.get("operations")
        if not operations or not isinstance(operations, list):
            return _err("excel.edit: 'operations' is required and must be a non-empty list")

        path = _resolve(raw)
        p = Path(path)
        if not p.exists():
            return _err(f"excel.edit: file not found: {path} — use fs_browse.find to locate it")

        try:
            import openpyxl
        except ImportError:
            return _err("openpyxl not installed. Run: pip install openpyxl")

        try:
            wb = openpyxl.load_workbook(str(p))
        except PermissionError:
            return _err(f"Permission denied: {path}")
        except Exception as e:
            return _err(f"Failed to open workbook: {e}")

        default_sheet: Optional[str] = args.get("sheet")
        results: List[Dict[str, Any]] = []

        for op in operations:
            if not isinstance(op, dict):
                results.append(_err(f"Each operation must be an object, got: {type(op).__name__}"))
                break
            result = self._apply_op(wb, default_sheet, op)
            results.append(result)
            if result.get("STATUS") == "failed":
                break

        all_ok = all(r.get("STATUS") == "success" for r in results)

        if all_ok:
            try:
                wb.save(str(p))
            except PermissionError:
                return _err(
                    f"Permission denied when saving: {path} — file may be open in another app.",
                    RESULTS=results,
                    SAVED=False,
                )
            except Exception as e:
                return _err(f"Operations succeeded but save failed: {e}", RESULTS=results, SAVED=False)

        out: Dict[str, Any] = {
            "STATUS": "success" if all_ok else "failed",
            "PATH": str(p),
            "OPERATIONS_TOTAL": len(operations),
            "OPERATIONS_APPLIED": len(results),
            "RESULTS": results,
            "SAVED": all_ok,
        }
        if not all_ok:
            failed = next((r for r in results if r.get("STATUS") == "failed"), {})
            out["ERROR"] = failed.get("ERROR") or "An operation failed — file was not modified."
        return out

    def _apply_op(
        self, wb: Any, default_sheet: Optional[str], op: Dict[str, Any]
    ) -> Dict[str, Any]:
        op_type = str(op.get("op") or "").strip().lower()

        cell_row_ops = {"set_cell", "add_row", "insert_row", "delete_row"}
        if op_type in cell_row_ops:
            sheet_name = op.get("sheet") or default_sheet
            if not sheet_name:
                return {
                    "op": op_type, "STATUS": "failed",
                    "ERROR": "sheet is required. Set it at the edit level or inside the op.",
                }
            if sheet_name not in wb.sheetnames:
                return {
                    "op": op_type, "STATUS": "failed",
                    "ERROR": f"Sheet '{sheet_name}' not found.",
                    "AVAILABLE_SHEETS": wb.sheetnames,
                }
            ws = wb[sheet_name]

        try:
            if op_type == "set_cell":
                cell = str(op.get("cell") or "").strip().upper()
                if not cell or not re.match(r"^[A-Z]+\d+$", cell):
                    return {
                        "op": op_type, "STATUS": "failed",
                        "ERROR": (
                            f"Invalid cell address: '{cell}'. "
                            "Use A1 notation: column letter(s) + row number. Examples: A1, B3, AA10."
                        ),
                    }
                value = op.get("value")
                ws[cell] = value
                return {"op": op_type, "STATUS": "success", "CELL": cell, "VALUE": value}

            elif op_type == "add_row":
                row = op.get("row")
                if row is None or not isinstance(row, list):
                    return {"op": op_type, "STATUS": "failed", "ERROR": "row must be a list of values"}
                ws.append(row)
                new_row_idx = ws.max_row
                
                # Inherit formatting from the row above
                if new_row_idx > 1:
                    from copy import copy
                    for col_idx in range(1, len(row) + 1):
                        new_cell = ws.cell(row=new_row_idx, column=col_idx)
                        prev_cell = ws.cell(row=new_row_idx - 1, column=col_idx)
                        if prev_cell.has_style:
                            new_cell.font = copy(prev_cell.font)
                            new_cell.border = copy(prev_cell.border)
                            new_cell.fill = copy(prev_cell.fill)
                            new_cell.number_format = copy(prev_cell.number_format)
                            new_cell.alignment = copy(prev_cell.alignment)

                return {
                    "op": op_type, "STATUS": "success",
                    "ROW_INDEX": new_row_idx, "COLUMNS_WRITTEN": len(row),
                }

            elif op_type == "insert_row":
                at_index = op.get("at_index")
                if at_index is None:
                    return {"op": op_type, "STATUS": "failed", "ERROR": "at_index is required (1-based)"}
                at_index = int(at_index)
                if at_index < 1:
                    return {"op": op_type, "STATUS": "failed", "ERROR": "at_index must be >= 1"}
                row = op.get("row") or []
                ws.insert_rows(at_index)
                
                from copy import copy
                for col_idx, val in enumerate(row, 1):
                    new_cell = ws.cell(row=at_index, column=col_idx, value=val)
                    # Inherit formatting from the row above (if exists) or the row that was pushed down
                    reference_row_idx = at_index - 1 if at_index > 1 else at_index + 1
                    ref_cell = ws.cell(row=reference_row_idx, column=col_idx)
                    
                    if ref_cell.has_style:
                        new_cell.font = copy(ref_cell.font)
                        new_cell.border = copy(ref_cell.border)
                        new_cell.fill = copy(ref_cell.fill)
                        new_cell.number_format = copy(ref_cell.number_format)
                        new_cell.alignment = copy(ref_cell.alignment)

                return {
                    "op": op_type, "STATUS": "success",
                    "INSERTED_AT": at_index, "COLUMNS_WRITTEN": len(row),
                }

            elif op_type == "delete_row":
                at_index = op.get("at_index")
                if at_index is None:
                    return {"op": op_type, "STATUS": "failed", "ERROR": "at_index is required (1-based)"}
                at_index = int(at_index)
                max_row = ws.max_row or 0
                if at_index < 1 or at_index > max_row:
                    return {
                        "op": op_type, "STATUS": "failed",
                        "ERROR": f"at_index {at_index} out of range. Sheet has {max_row} row(s).",
                    }
                ws.delete_rows(at_index)
                return {"op": op_type, "STATUS": "success", "DELETED_ROW": at_index}

            elif op_type == "add_sheet":
                name = str(op.get("name") or "").strip()
                if not name:
                    return {"op": op_type, "STATUS": "failed", "ERROR": "name is required"}
                if name in wb.sheetnames:
                    return {
                        "op": op_type, "STATUS": "failed",
                        "ERROR": f"Sheet '{name}' already exists.",
                        "EXISTING_SHEETS": wb.sheetnames,
                    }
                new_ws = wb.create_sheet(name)
                headers = op.get("headers") or []
                if headers:
                    new_ws.append([str(h) for h in headers])
                return {
                    "op": op_type, "STATUS": "success",
                    "SHEET": name, "HEADERS_WRITTEN": len(headers),
                }

            elif op_type == "delete_sheet":
                name = str(op.get("name") or "").strip()
                if not name:
                    return {"op": op_type, "STATUS": "failed", "ERROR": "name is required"}
                if name not in wb.sheetnames:
                    return {
                        "op": op_type, "STATUS": "failed",
                        "ERROR": f"Sheet '{name}' not found.",
                        "AVAILABLE_SHEETS": wb.sheetnames,
                    }
                if not op.get("confirmed"):
                    return {
                        "op": op_type, "STATUS": "failed",
                        "NEEDS_CONFIRMATION": True,
                        "PREVIEW": f"Will permanently delete sheet '{name}' and all its data.",
                        "NOTE": 'Add "confirmed": true inside this op object to proceed.',
                    }
                del wb[name]
                return {"op": op_type, "STATUS": "success", "DELETED_SHEET": name}

            elif op_type == "rename_sheet":
                name = str(op.get("name") or "").strip()
                new_name = str(op.get("new_name") or "").strip()
                if not name or not new_name:
                    return {
                        "op": op_type, "STATUS": "failed",
                        "ERROR": "name and new_name are both required",
                    }
                if name not in wb.sheetnames:
                    return {
                        "op": op_type, "STATUS": "failed",
                        "ERROR": f"Sheet '{name}' not found.",
                        "AVAILABLE_SHEETS": wb.sheetnames,
                    }
                if new_name in wb.sheetnames:
                    return {
                        "op": op_type, "STATUS": "failed",
                        "ERROR": f"Sheet '{new_name}' already exists. Choose a different new_name.",
                    }
                wb[name].title = new_name
                return {"op": op_type, "STATUS": "success", "OLD_NAME": name, "NEW_NAME": new_name}

            else:
                valid = "set_cell, add_row, insert_row, delete_row, add_sheet, delete_sheet, rename_sheet"
                return {
                    "op": op_type or "(empty)", "STATUS": "failed",
                    "ERROR": f"Unknown op: '{op_type}'. Valid ops: {valid}",
                }

        except Exception as e:
            return {"op": op_type, "STATUS": "failed", "ERROR": str(e)}

    # ── search ───────────────────────────────────────────────────────────────

    def _search(self, args: Dict[str, Any]) -> Dict[str, Any]:
        raw = str(args.get("path") or "").strip()
        if not raw:
            return _err("excel.search: 'path' is required — provide an absolute .xlsx path")

        path = _resolve(raw)
        p = Path(path)
        if not p.exists():
            return _err(f"excel.search: file not found: {path} — use fs_browse.find to locate it")

        sheet_name: Optional[str] = args.get("sheet")
        query = args.get("query")
        operator = args.get("operator")

        if query is None and operator is None:
            return _err(
                "Provide either query (text search) or column + operator + value (condition search)"
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), data_only=True)

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return _err(f"Sheet '{sheet_name}' not found.", AVAILABLE_SHEETS=wb.sheetnames)
                ws = wb[sheet_name]
            else:
                ws = wb.active

            if query is not None:
                return self._search_text(ws, str(query), args.get("column"))

            column = args.get("column")
            value = args.get("value")
            if not column:
                return _err("column is required for condition search")
            if value is None:
                return _err("value is required for condition search")
            return self._search_condition(ws, str(column), str(operator), value)

        except ImportError:
            return _err("openpyxl not installed. Run: pip install openpyxl")
        except PermissionError:
            return _err(f"Permission denied: {path}")
        except Exception as e:
            return _err(f"Search failed: {e}")

    def _search_text(
        self, ws: Any, query: str, column: Optional[str]
    ) -> Dict[str, Any]:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return _ok(SHEET=ws.title, QUERY=query, MATCHES=[], TOTAL=0)

        headers = [str(h) if h is not None else "" for h in all_rows[0]]
        headers_lower = [h.lower() for h in headers]
        query_lower = query.lower()

        col_idx: Optional[int] = None
        if column:
            col_key = column.lower()
            if col_key not in headers_lower:
                return _err(f"Column '{column}' not found.", AVAILABLE_COLUMNS=headers)
            col_idx = headers_lower.index(col_key)

        matches: List[Dict[str, Any]] = []
        for row_idx, row in enumerate(all_rows[1:], 2):
            if col_idx is not None:
                val = row[col_idx] if col_idx < len(row) else None
                if query_lower in str(val if val is not None else "").lower():
                    matches.append({
                        "row": row_idx,
                        "cell": f"{_col_letter(col_idx + 1)}{row_idx}",
                        "column": headers[col_idx],
                        "value": val,
                        "full_row": dict(zip(headers, row)),
                    })
            else:
                for c_idx, val in enumerate(row):
                    if query_lower in str(val if val is not None else "").lower():
                        matches.append({
                            "row": row_idx,
                            "cell": f"{_col_letter(c_idx + 1)}{row_idx}",
                            "column": headers[c_idx] if c_idx < len(headers) else f"Col{c_idx + 1}",
                            "value": val,
                            "full_row": dict(zip(headers, row)),
                        })
                        break  # one match per row

        return _ok(
            SHEET=ws.title, QUERY=query,
            COLUMN_FILTER=column, MATCHES=matches, TOTAL=len(matches),
        )

    def _search_condition(
        self, ws: Any, column: str, operator: str, value: Any
    ) -> Dict[str, Any]:
        valid_ops = {"=", "!=", ">", "<", ">=", "<=", "contains", "starts_with", "ends_with"}
        if operator not in valid_ops:
            return _err(
                f"Invalid operator: '{operator}'. Valid: {', '.join(sorted(valid_ops))}"
            )

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return _ok(SHEET=ws.title, COLUMN=column, OPERATOR=operator, VALUE=value,
                       MATCHES=[], TOTAL=0)

        headers = [str(h) if h is not None else "" for h in all_rows[0]]
        headers_lower = [h.lower() for h in headers]

        if column.lower() not in headers_lower:
            return _err(f"Column '{column}' not found.", AVAILABLE_COLUMNS=headers)
        col_idx = headers_lower.index(column.lower())

        matches: List[Dict[str, Any]] = []
        for row_idx, row in enumerate(all_rows[1:], 2):
            cell_val = row[col_idx] if col_idx < len(row) else None
            if self._match_condition(cell_val, operator, value):
                matches.append({
                    "row": row_idx,
                    "cell": f"{_col_letter(col_idx + 1)}{row_idx}",
                    "value": cell_val,
                    "full_row": dict(zip(headers, row)),
                })

        return _ok(
            SHEET=ws.title, COLUMN=column, OPERATOR=operator, VALUE=value,
            MATCHES=matches, TOTAL=len(matches),
        )

    def _match_condition(self, cell_val: Any, operator: str, value: Any) -> bool:
        try:
            if operator in ("contains", "starts_with", "ends_with"):
                s = str(cell_val).lower() if cell_val is not None else ""
                v = str(value).lower()
                if operator == "contains":
                    return v in s
                if operator == "starts_with":
                    return s.startswith(v)
                return s.endswith(v)

            if operator in (">", "<", ">=", "<="):
                lhs, rhs = float(cell_val), float(value)
                if operator == ">":
                    return lhs > rhs
                if operator == "<":
                    return lhs < rhs
                if operator == ">=":
                    return lhs >= rhs
                return lhs <= rhs

            # = and !=
            if isinstance(value, str):
                match = (
                    str(cell_val).lower() == value.lower()
                    if cell_val is not None
                    else value.lower() == ""
                )
            else:
                match = cell_val == value
            return match if operator == "=" else not match

        except (TypeError, ValueError):
            return False



# ── registry ─────────────────────────────────────────────────────────────────

TOOL_NAME = "excel"
TOOL_CLASS = ExcelTool


def get_tool() -> ExcelTool:
    return ExcelTool()
